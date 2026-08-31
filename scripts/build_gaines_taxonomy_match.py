import sys
import openpyxl
from openpyxl.utils import column_index_from_string

sys.path.insert(0, "scripts")
from matching_utils import top_matches_batch

GAINES_XLSX = "/root/.claude/uploads/9c3a0553-6723-59bf-af99-2ae863779397/3a2ebc01-Gaines_US_EPA_PFA_Mapping.xlsx"
ECHA_XLSX = "/root/.claude/uploads/9c3a0553-6723-59bf-af99-2ae863779397/55d8c115-ECHA_compiled.xlsx"
OUT_XLSX = "output/gaines_taxonomy_match.xlsx"

wb_g = openpyxl.load_workbook(GAINES_XLSX, data_only=True)
ws_g = wb_g["Table 1"]
assert ws_g.cell(row=1, column=5).value == "Details and references"
details = []
for r in range(2, ws_g.max_row + 1):
    v = ws_g.cell(row=r, column=5).value
    if v is not None and str(v).strip() != "":
        details.append(str(v).strip())
details = sorted(set(details))
print("unique 'Details and references' values:", len(details))

wb_e = openpyxl.load_workbook(ECHA_XLSX, data_only=True)
ws_e = wb_e["Downstream Use Taxonomy"]
col_as = column_index_from_string("AS")
as_values = []
for r in range(7, ws_e.max_row + 1):
    v = ws_e.cell(row=r, column=col_as).value
    if v is not None and str(v).strip() != "":
        as_values.append(str(v).strip())
as_values = sorted(set(as_values))
print("unique AS values:", len(as_values))

TOP_N = 3
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Gaines-Taxonomy Match"
headers = ["Gaines Table 1, column E (Details and references)"]
for i in range(1, TOP_N + 1):
    headers += [f"best match {i} (Downstream Use Taxonomy, col AS)", f"similarity {i} (0-100)"]
ws_out.append(headers)

all_matches = top_matches_batch(details, as_values, top_n=TOP_N)
for d, matches in zip(details, all_matches):
    line = [d]
    for match_text, score in matches:
        line += [match_text, score]
    ws_out.append(line)

widths = [70] + [55, 12] * TOP_N
for i, w in enumerate(widths, start=1):
    ws_out.column_dimensions[ws_out.cell(row=1, column=i).column_letter].width = w
ws_out.freeze_panes = "A2"

notes = wb_out.create_sheet("Notes")
notes.append(["Method"])
notes.append(["For each unique value in Gaines_US_EPA_PFA_Mapping.xlsx, sheet 'Table 1', column E ('Details and references' -- a free-text"])
notes.append(["use/application description per chemical row), the top 3 closest text matches were found among the unique values in"])
notes.append(["ECHA_compiled.xlsx, sheet 'Downstream Use Taxonomy', column AS ('PFAS-reliant product/component'), using RapidFuzz's"])
notes.append(["token_set_ratio (0-100; 100 = same words, order/case-insensitive)."])
notes.append(["Before scoring, both sides are expanded with a domain synonym dictionary (scripts/matching_utils.py) so e.g. a clothing-item name"])
notes.append(["matches a general 'clothing/apparel' taxonomy entry, and PFAS-chemistry abbreviations match their spelled-out form."])
notes.append([""])
notes.append(["A high score means the two text strings share most of their words; it is NOT a chemical or regulatory equivalence claim."])
notes.append(["Scores below ~50 are generally weak matches and should be treated as 'no good match found' -- spot-check before relying on them."])
notes.column_dimensions["A"].width = 120

wb_out.save(OUT_XLSX)
print("wrote", OUT_XLSX)
