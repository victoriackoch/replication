import sys
import openpyxl
from openpyxl.utils import column_index_from_string

sys.path.insert(0, "scripts")
from matching_utils import top_matches_batch

CG_XLSX = "/root/.claude/uploads/9c3a0553-6723-59bf-af99-2ae863779397/c777292b-US_Consumer_Goods_Goods_only.xlsx"
ECHA_XLSX = "/root/.claude/uploads/9c3a0553-6723-59bf-af99-2ae863779397/55d8c115-ECHA_compiled.xlsx"
OUT_XLSX = "output/consumer_goods_taxonomy_gap_analysis.xlsx"
GAP_THRESHOLD = 45  # token_set_ratio below this = "likely missing from column AS"

wb_cg = openpyxl.load_workbook(CG_XLSX, data_only=True)
ws_cg = wb_cg["Sources in Consumer Products"]
assert ws_cg.cell(row=4, column=9).value == "Product"
products = []
for r in range(5, ws_cg.max_row + 1):
    v = ws_cg.cell(row=r, column=9).value
    if v is not None and str(v).strip() != "":
        products.append(str(v).strip())
products = sorted(set(products))
print("unique column-I products:", len(products))

wb_e = openpyxl.load_workbook(ECHA_XLSX, data_only=True)
ws_e = wb_e["Downstream Use Taxonomy"]
col_as = column_index_from_string("AS")
as_values = []
for r in range(7, ws_e.max_row + 1):
    v = ws_e.cell(row=r, column=col_as).value
    if v is not None and str(v).strip() != "":
        as_values.append(str(v).strip())
as_values = sorted(set(as_values))
print("unique AS values:", len(as_values))

TOP_N = 3
all_matches = top_matches_batch(products, as_values, top_n=TOP_N)
rows = []
for p, matches in zip(products, all_matches):
    best_score = matches[0][1] if matches else 0
    rows.append((p, matches, best_score))

# sort weakest matches first, so likely gaps surface at the top
rows.sort(key=lambda x: x[2])

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Gap analysis (all products)"
headers = ["Consumer Products, column I (Product)", "likely missing from col AS?"]
for i in range(1, TOP_N + 1):
    headers += [f"best match {i} (Downstream Use Taxonomy, col AS)", f"similarity {i} (0-100)"]
ws_out.append(headers)
for p, matches, best_score in rows:
    line = [p, "YES" if best_score < GAP_THRESHOLD else ""]
    for match_text, score in matches:
        line += [match_text, score]
    ws_out.append(line)
widths = [55, 16] + [55, 12] * TOP_N
for i, w in enumerate(widths, start=1):
    ws_out.column_dimensions[ws_out.cell(row=1, column=i).column_letter].width = w
ws_out.freeze_panes = "A2"

gap_only = wb_out.create_sheet("Likely gaps only")
gap_only.append(["Consumer Products, column I (Product)", "best match found (col AS)", "similarity (0-100)"])
n_gaps = 0
for p, matches, best_score in rows:
    if best_score < GAP_THRESHOLD:
        gap_only.append([p, matches[0][0] if matches else "", best_score])
        n_gaps += 1
gap_only.column_dimensions["A"].width = 55
gap_only.column_dimensions["B"].width = 55
gap_only.column_dimensions["C"].width = 14
gap_only.freeze_panes = "A2"
print("likely-gap products (score <", GAP_THRESHOLD, "):", n_gaps, "of", len(products))

notes = wb_out.create_sheet("Notes")
notes.append(["Method"])
notes.append(["For each unique value in US_Consumer_Goods_Goods_only.xlsx, sheet 'Sources in Consumer Products', column I (Product),"])
notes.append(["the top 3 closest text matches were found among the unique values in ECHA_compiled.xlsx, sheet 'Downstream Use Taxonomy',"])
notes.append(["column AS ('PFAS-reliant product/component'), using RapidFuzz's token_set_ratio (0-100; 100 = same words, order/case-insensitive)."])
notes.append(["Before scoring, both sides are expanded with a domain synonym dictionary (scripts/matching_utils.py) so e.g. a clothing-item name"])
notes.append(["matches a general 'clothing/apparel' taxonomy entry, and PFAS-chemistry abbreviations match their spelled-out form."])
notes.append([f"A product is flagged 'likely missing from col AS' when its best match scores below {GAP_THRESHOLD} -- i.e. the product name shares"])
notes.append(["little or no wording with anything already in that taxonomy column, suggesting that kind of product/component isn't represented there yet."])
notes.append([""])
notes.append([f"{n_gaps} of {len(products)} unique products fall below the {GAP_THRESHOLD} threshold -- see the 'Likely gaps only' sheet."])
notes.append(["This is a text-similarity heuristic, not a verified taxonomy audit: a low score can also mean the product is phrased very differently"])
notes.append(["from the taxonomy's wording despite being conceptually covered, and a moderate score does not guarantee real overlap. Spot-check before relying on it."])
notes.column_dimensions["A"].width = 120

wb_out.save(OUT_XLSX)
print("wrote", OUT_XLSX)
