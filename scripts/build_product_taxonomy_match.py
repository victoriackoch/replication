import csv
import openpyxl
from openpyxl.utils import column_index_from_string
from rapidfuzz import fuzz, process

MAIN_CSV = "output/pfas_product_substance_table.csv"
ECHA_XLSX = "/root/.claude/uploads/9c3a0553-6723-59bf-af99-2ae863779397/55d8c115-ECHA_compiled.xlsx"
OUT_XLSX = "output/pfas_product_taxonomy_match.xlsx"

with open(MAIN_CSV) as f:
    rows = list(csv.DictReader(f))
products = sorted(set(r["product"].strip() for r in rows if r["product"] and r["product"].strip()))
print("unique products:", len(products))

wb_in = openpyxl.load_workbook(ECHA_XLSX, data_only=True)
ws = wb_in["Downstream Use Taxonomy"]
col_as = column_index_from_string("AS")
as_values = []
for r in range(7, ws.max_row + 1):
    v = ws.cell(row=r, column=col_as).value
    if v is not None and str(v).strip() != "":
        as_values.append(str(v).strip())
as_values = sorted(set(as_values))
print("unique AS values:", len(as_values))

TOP_N = 3
results = []
for p in products:
    matches = process.extract(p, as_values, scorer=fuzz.token_set_ratio, limit=TOP_N)
    row = {"product": p}
    for i, (match_text, score, _) in enumerate(matches, start=1):
        row[f"match_{i}"] = match_text
        row[f"score_{i}"] = round(score, 1)
    results.append(row)

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Product-Taxonomy Match"
headers = ["product (main table, column A)"]
for i in range(1, TOP_N + 1):
    headers += [f"best match {i} (Downstream Use Taxonomy, col AS)", f"similarity {i} (0-100)"]
ws_out.append(headers)
for row in results:
    line = [row["product"]]
    for i in range(1, TOP_N + 1):
        line.append(row.get(f"match_{i}", ""))
        line.append(row.get(f"score_{i}", ""))
    ws_out.append(line)

# widths
widths = [55] + [55, 12] * TOP_N
for i, w in enumerate(widths, start=1):
    ws_out.column_dimensions[ws_out.cell(row=1, column=i).column_letter].width = w
ws_out.freeze_panes = "A2"

notes = wb_out.create_sheet("Notes")
notes.append(["Method"])
notes.append(["For each unique 'product' value in the main product/substance table (output/pfas_product_substance_table.csv, column A),"])
notes.append(["the top 3 closest text matches were found among the unique values in ECHA_compiled.xlsx, sheet 'Downstream Use Taxonomy', column AS"])
notes.append(["('PFAS-reliant product/component', per that sheet's row-6 header), using RapidFuzz's token_set_ratio (0-100; 100 = same words, order/case-insensitive)."])
notes.append([""])
notes.append(["A high score means the two text strings share most of their words; it is NOT a chemical or regulatory equivalence claim."])
notes.append(["Scores below ~50 are generally weak matches (different topic) and should be treated as 'no good match found' -- spot-check before relying on them."])
notes.column_dimensions["A"].width = 120

wb_out.save(OUT_XLSX)
print("wrote", OUT_XLSX)
