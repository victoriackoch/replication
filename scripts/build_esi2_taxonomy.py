import re
import openpyxl

ESI2_XLSX = "/root/.claude/uploads/9c3a0553-6723-59bf-af99-2ae863779397/7c12dc56-ESI2.xlsm"
OUT_XLSX = "output/pfas_esi2_taxonomy_table.xlsx"

DATA_SHEETS = [
    "PFAAs", "PFPIA-based", "PASF-based", "PACF-based", "Fluorotelomer-based",
    "Cyclic PFAS", "Other Nonpolymers", "Side-chain fluorinated aromatic",
    "Per- and polyfluoroalkyl ether ", "Hydrofluoroether", "Non-polymers_Polymers",
    "Fluoropolymer", "Side-chain fluorinated polymers", "Perfluoropolyether",
]

# sheet name -> the group label used on the "Overview" sheet
GROUP_LABEL = {
    "PFAAs": "Perfluoroalkyl acids (PFAAs)",
    "PFPIA-based": "Perfluoroalkyl phosphinic acids (PFPIA)-based substances",
    "PASF-based": "Perfluoroalkane sulfonyl fluoride (PASF)-based substances",
    "PACF-based": "Perfluoroalkyl carbonyl fluoride (PACF)-based substances",
    "Fluorotelomer-based": "Fluorotelomer-based substances",
    "Cyclic PFAS": "Cyclic PFAS",
    "Other Nonpolymers": "Other Nonpolymers",
    "Side-chain fluorinated aromatic": "Aromatics with fluorinated side-chains",
    "Per- and polyfluoroalkyl ether ": "Per- and polyfluoroalkyl ether",
    "Hydrofluoroether": "Hydofluoroether",
    "Non-polymers_Polymers": "Non-polymers? Polymers?",
    "Fluoropolymer": "Fluoropolymers",
    "Side-chain fluorinated polymers": "Side-chain fluorinated polymers",
    "Perfluoropolyether": "Perfluoropolyether",
}

NAME_PAREN_RE = re.compile(r"^(.*?)\s*\(([^()]+)\)\s*$")
# an abbreviation is a short, mostly-uppercase single "word" (letters,
# digits, hyphens) -- guards against splitting a trailing parenthetical
# that's actually part of the chemical name itself, e.g. "Poly(vinylidene
# fluoride)" or "Perfluoro(propyl vinyl ether)", where the parenthesised
# text is a lowercase, multi-word descriptive phrase, not an acronym.
ABBREV_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9\-]{0,9}$")


def looks_like_abbrev(token):
    return bool(ABBREV_RE.match(token)) and any(c.isupper() for c in token)


def split_name_abbrev(name):
    if not name:
        return name, None
    name = str(name).strip()
    m = NAME_PAREN_RE.match(name)
    if m and looks_like_abbrev(m.group(2).strip()):
        return m.group(1).strip(), m.group(2).strip()
    return name, None


# A sub-group/sub-sub-group header row's own text is "<name>   <generic
# structure>" (point 3): the name and the class's generic formula
# (CnF2n+1..., often with a counter-ion) separated visually by a run of
# 2+ spaces. Not every row uses a double space, though ("...(PFSAs)
# CnF2n+1SO3H" has only one) -- the fallback looks for the last single
# space directly before a token shaped like a formula (starts with a
# capital-C element/counter symbol followed by a lowercase "n", another
# capital letter, a digit, or "(").
DOUBLE_SPACE_RE = re.compile(r"\s{2,}")
FORMULA_TAIL_RE = re.compile(r"\s([A-Z][a-zA-Z0-9(].*)$")


def split_name_structure(text):
    if not text:
        return text, None
    text = str(text).strip()
    m = DOUBLE_SPACE_RE.search(text)
    if m:
        return text[:m.start()].rstrip(), text[m.end():].strip()
    m = FORMULA_TAIL_RE.search(text)
    if m and re.search(r"[A-Z].*[a-z]n[A-Z0-9]|[0-9]", m.group(1)):
        return text[:m.start()].rstrip(), m.group(1).strip()
    return text, None


# Header-name -> logical field, matched by normalized (lowercased,
# whitespace-collapsed) column-1 header text. Column layout is NOT
# consistent across sheets (Fluoropolymer inserts an extra "Acronym 2"
# column and drops "Structure"; Non-polymers_Polymers drops the mass
# column entirely) -- resolving every column by its own sheet's header
# row, rather than a fixed position shared across sheets, is the fix for
# both: reading fixed positions previously walked off into the wrong
# column on both of those sheets (values from "Reference"/"Function"
# ending up in the cas_number/oecd_inclusion fields).
FIELD_MATCHERS = [
    ("acronym", lambda h: h in ("acronym", "acronym 1")),
    ("acronym_2", lambda h: h == "acronym 2"),
    ("name", lambda h: h == "name"),
    ("structure", lambda h: h == "structure"),
    ("chemical_formula", lambda h: h == "chemical formula"),
    ("mass", lambda h: h.startswith("mono-isotopic mass")),
    ("cas", lambda h: h == "cas no."),
    ("trade_names", lambda h: h.startswith("trade name")),
    ("oecd", lambda h: h.startswith("included in oecd")),
]


def normalize_header(h):
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).strip().lower())


def build_column_map(ws):
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = normalize_header(ws.cell(row=1, column=c).value)
        if not h:
            continue
        for field, matcher in FIELD_MATCHERS:
            if field in col_map:
                continue
            if matcher(h):
                col_map[field] = c
                break
    return col_map


def fill_signature(cell):
    f = cell.fill
    if f is None or f.fgColor is None or f.fill_type is None:
        return ("none", None, None)
    try:
        theme = f.fgColor.theme
    except Exception:
        theme = None
    try:
        tint = round(f.fgColor.tint, 3) if theme is not None else None
    except Exception:
        tint = None
    return (f.fill_type, theme, tint)


def is_subgroup_row(sig):
    return sig[0] == "solid" and sig[1] == 1


def is_subsubgroup_row(sig):
    return sig[0] == "solid" and sig[1] == 2


def cell_str(v):
    return str(v).strip() if v is not None else None


wb_in = openpyxl.load_workbook(ESI2_XLSX, data_only=True)

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "ESI-2 Taxonomy"
headers = ["group", "sub_group", "sub_sub_group_name", "sub_sub_group_structure",
           "acronym", "acronym_2", "name", "abbreviation", "chemical_formula",
           "mono_isotopic_mass", "cas_number", "trade_names", "oecd_inclusion"]
ws_out.append(headers)

total = 0
for sheet in DATA_SHEETS:
    ws = wb_in[sheet]
    group = GROUP_LABEL.get(sheet, sheet)
    col_map = build_column_map(ws)
    assert "name" in col_map and "cas" in col_map, f"{sheet}: could not resolve Name/CAS columns from headers"

    current_subgroup = None
    current_subsub_name = None
    current_subsub_structure = None

    for r in range(2, ws.max_row + 1):
        sig = fill_signature(ws.cell(row=r, column=1))
        label = cell_str(ws.cell(row=r, column=1).value)

        if is_subgroup_row(sig):
            if label:
                current_subgroup = label
            current_subsub_name = None
            current_subsub_structure = None
            continue
        if is_subsubgroup_row(sig):
            if label:
                current_subsub_name, current_subsub_structure = split_name_structure(label)
            continue

        name_raw = ws.cell(row=r, column=col_map["name"]).value
        formula = ws.cell(row=r, column=col_map["chemical_formula"]).value if "chemical_formula" in col_map else None
        has_name = name_raw is not None and str(name_raw).strip() != ""
        has_formula = formula is not None and str(formula).strip() != ""
        if not has_name and not has_formula:
            continue  # blank separator / image-placeholder row

        name, abbrev = split_name_abbrev(name_raw) if has_name else (None, None)
        acronym = cell_str(ws.cell(row=r, column=col_map["acronym"]).value) if "acronym" in col_map else None
        acronym_2 = cell_str(ws.cell(row=r, column=col_map["acronym_2"]).value) if "acronym_2" in col_map else None
        mass = ws.cell(row=r, column=col_map["mass"]).value if "mass" in col_map else None
        cas = cell_str(ws.cell(row=r, column=col_map["cas"]).value)
        trade_names = ws.cell(row=r, column=col_map["trade_names"]).value if "trade_names" in col_map else None
        oecd = ws.cell(row=r, column=col_map["oecd"]).value if "oecd" in col_map else None

        ws_out.append([
            group, current_subgroup, current_subsub_name, current_subsub_structure,
            acronym, acronym_2, name, abbrev, formula, mass, cas, trade_names, oecd,
        ])
        total += 1

widths = [30, 40, 40, 30, 12, 14, 45, 16, 20, 14, 16, 25, 16]
for i, w in enumerate(widths, start=1):
    ws_out.column_dimensions[ws_out.cell(row=1, column=i).column_letter].width = w
ws_out.freeze_panes = "A2"

notes = wb_out.create_sheet("Notes")
notes.append(["Built from ESI2.xlsm."])
notes.append(["'group' = the substance-family sheet the row came from (matches the 'Group' column on the 'Overview' sheet)."])
notes.append(["'sub_group' = the text of the dark-grey header row above the row's block (e.g. 'PFCAs', 'Fluoropolymers - monoconstituents'),"])
notes.append(["identified by that row's actual cell fill (theme colour 1), not by position -- it stays in effect until the next dark-grey row."])
notes.append(["'sub_sub_group_name' / 'sub_sub_group_structure' = a second, differently-shaded header row that sometimes sits between the"])
notes.append(["sub_group row and the data (e.g. 'Perfluoroalkyl carboxylic acids (PFCAs)' / 'CnF2n+1COOH'), identified by its own fill (theme"])
notes.append(["colour 2) and split on the run of 2+ spaces between the class name and its generic formula (a single-space fallback catches the"])
notes.append(["handful of rows that don't use a double space). Blank when no such row precedes the block."])
notes.append(["'name'/'abbreviation' = the sheet's own 'Name' column split the same way on a single trailing parenthetical, e.g."])
notes.append(["'Perfluorobutanoic acid (PFBA)' -> name 'Perfluorobutanoic acid', abbreviation 'PFBA'."])
notes.append(["'acronym'/'acronym_2' = the sheet's 'Acronym' (or 'Acronym 1') column and, only on the Fluoropolymer sheet, its second"])
notes.append(["'Acronym 2' column (a monomer-composition shorthand like 'TFE-HFP' alongside the polymer's own acronym like 'FEP')."])
notes.append([""])
notes.append(["Every other column (chemical_formula, mono_isotopic_mass, cas_number, trade_names, oecd_inclusion) is read from whichever"])
notes.append(["column that sheet's own row-1 header names it, rather than a fixed position -- two sheets use a different column layout"])
notes.append(["(Fluoropolymer inserts 'Acronym 2' and drops 'Structure'; Non-polymers_Polymers drops the mass column entirely), which a"])
notes.append(["fixed-position read had been silently misaligning (pulling Reference-CAS/Function text into the cas_number/oecd_inclusion"])
notes.append(["columns, and skipping nearly every Fluoropolymer row because it mistook the blank 'Acronym 2' column for the 'Name' column)."])
notes.column_dimensions["A"].width = 118

wb_out.save(OUT_XLSX)
print("rows:", total)
print("wrote", OUT_XLSX)
