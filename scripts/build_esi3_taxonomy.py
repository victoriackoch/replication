import openpyxl

ESI3_XLSX = "/root/.claude/uploads/9c3a0553-6723-59bf-af99-2ae863779397/5741f8f6-ESI3.xlsm"
OUT_XLSX = "output/pfas_esi3_taxonomy_table.xlsx"

wb_in = openpyxl.load_workbook(ESI3_XLSX, data_only=True)
sheets = [s for s in wb_in.sheetnames if s != "Explanations"]

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "ESI-3 Taxonomy"
ws_out.append(["sheet_name", "header", "cas_number", "name", "elemental_composition", "use_type", "pfas_type"])

total = 0
for sheet in sheets:
    ws = wb_in[sheet]
    current_header = None
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        a, b, c, d, e, f = row
        if a is None and b is None and c is None and d is None and e is None and f is None:
            continue  # blank separator row
        if a == "CAS No." and b == "Name":
            continue  # literal column-header row, repeats per block
        if b is None and c is None and d is None and e is None and f is None:
            # sub-header row: only column A has a value
            if a is not None and str(a).strip() != "":
                current_header = str(a).strip()
            continue
        # data row
        ws_out.append([
            sheet, current_header,
            str(a).strip() if a is not None else None,
            b, c, d, e,
        ])
        total += 1

widths = [26, 34, 16, 55, 22, 10, 12]
for i, w in enumerate(widths, start=1):
    ws_out.column_dimensions[ws_out.cell(row=1, column=i).column_letter].width = w
ws_out.freeze_panes = "A2"

notes = wb_out.create_sheet("Notes")
notes.append(["Built from ESI3.xlsm."])
notes.append(["'sheet_name' = the source workbook tab (a use-sector or application-type category)."])
notes.append(["'header' = the sub-category label printed above that row's data block on its sheet (e.g. 'Brake and hydraulic fluids' on the Aerospace sheet),"])
notes.append(["forward-filled down through the block's data rows; each sheet has several such blocks."])
notes.append(["The remaining columns are each block's own table columns (CAS No./Name/Elemental composition/Use Type/PFAS Type) verbatim;"])
notes.append(["the 'Explanations' sheet (which defines Use Type and PFAS Type codes) was not a data sheet and is not included in the rows above -- see it in the source workbook."])
notes.column_dimensions["A"].width = 110

wb_out.save(OUT_XLSX)
print("rows:", total)
print("wrote", OUT_XLSX)
