import csv
import openpyxl

OLD_MATCH_XLSX = "/root/.claude/uploads/9c3a0553-6723-59bf-af99-2ae863779397/739fc4ef-pfas_product_taxonomy_match_2.xlsx"
MAIN_CSV = "output/pfas_product_substance_table.csv"
OUT_XLSX = "output/pfas_product_taxonomy_match.xlsx"

wb_old = openpyxl.load_workbook(OLD_MATCH_XLSX, data_only=True)
ws_old = wb_old["Product-Taxonomy Match"]

old_kept = {}
for r in range(2, ws_old.max_row + 1):
    product = ws_old.cell(row=r, column=1).value
    if not product:
        continue
    product = str(product).strip()
    matches = []
    for c in (2, 4, 6):  # best_match_1, best_match_2, best_match_3 slots
        v = ws_old.cell(row=r, column=c).value
        if v and str(v).strip():
            v = str(v).strip()
            if v not in matches:
                matches.append(v)
    if matches:
        old_kept[product] = matches[:2]  # at most 2 (columns B and C)

with open(MAIN_CSV) as f:
    rows = list(csv.DictReader(f))
new_products = sorted(set(r["product"].strip() for r in rows if r["product"] and r["product"].strip()))

print("unique products (new):", len(new_products))
print("products with a carried-forward kept match:", sum(1 for p in new_products if p in old_kept))
not_found = set(old_kept) - set(new_products)
print("previously-kept products no longer in the product list:", len(not_found))
for p in sorted(not_found):
    print("  DROPPED:", p)

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Product-Taxonomy Match"
ws_out.append([
    "product (main table, column A)",
    "kept match 1 (Downstream Use Taxonomy, col AS)",
    "kept match 2 (Downstream Use Taxonomy, col AS)",
])
for p in new_products:
    kept = old_kept.get(p, [])
    line = [p] + kept + [""] * (2 - len(kept))
    ws_out.append(line)

widths = [70, 55, 55]
for i, w in enumerate(widths, start=1):
    ws_out.column_dimensions[ws_out.cell(row=1, column=i).column_letter].width = w
ws_out.freeze_panes = "A2"

notes = wb_out.create_sheet("Notes")
notes.append(["This file was rebuilt after Table A.105 (food-contact/packaging PFAS) was manually re-derived to use the"])
notes.append(["'Use - taking into account function' column for the product name, which changed A.105's product strings."])
notes.append(["Column A (product) reflects the current, regenerated product list from output/pfas_product_substance_table.csv."])
notes.append(["Columns B and C carry forward the manually-kept match(es) from the previous review pass"])
notes.append(["(pfas_product_taxonomy_match_2.xlsx), re-associated to the same product text -- none of the 192 previously-kept"])
notes.append(["matches were on an A.105-derived product, so all 192 carried over unchanged with no re-association needed."])
notes.append(["Products with no manual judgement yet (including any newly-worded A.105 products) are left blank in B/C --"])
notes.append(["no new automated suggestions were generated for them; this pass only refreshed column A and reattached prior work."])
notes.column_dimensions["A"].width = 120

wb_out.save(OUT_XLSX)
print("wrote", OUT_XLSX)
